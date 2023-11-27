from icecream import ic
import openpyxl
import os
import re

path = r"F:\Temp\NIAC"
file_name = r"quotes_6_68_25-10-2023.xlsx"
full_path = os.path.join(path, file_name)

output_file = r"F:\Temp\NIAC\r_quotes_6_68_25-10-2023.xlsx"

ic(full_path)
sheet_name = "data"


wb = openpyxl.load_workbook(filename=full_path)
ws = wb[sheet_name]
ic(ws.max_row)
ic(ws.max_column)

column_f = 6

table_re = re.compile(r"\d+\.\d+(-\d+){4}")
wrong_code = re.compile(r"\d+\.\d+(-\d+){3}") # 6.51-1-0-2

current_table = ""

for row in range(1, ws.max_row):
    # ic(row, ws.cell(row, column_f).value)
    value = ws.cell(row, column_f).value
    code = "" if value is None else str(value)

    if table_re.match(code):
        current_table = code
        ic(code)
        continue

    if current_table and wrong_code.match(code):
        ic(current_table, code)
        d = ws.cell(row=row, column=column_f, value=current_table)


else:
    mess = f"пройдено {row} строк. "
    ic(mess)


wb.save(output_file)
ic("готово!")

