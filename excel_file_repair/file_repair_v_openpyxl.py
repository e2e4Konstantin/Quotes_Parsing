from icecream import ic
import openpyxl
import os
import re

# path = r"F:\Kazak\GoogleDrive\NIAC\parameterisation"
path = r"F:\Kazak\GoogleDrive\NIAC\Велесь_Игорь\Файлы к переводу в машиночитаемый вид\07-12-2023"

file_name = r"Расценки_5_68.xlsx"
full_path = os.path.join(path, file_name)

output_file_name = f"rep_{file_name}"
output_file = os.path.join(path, output_file_name)

ic(full_path)
sheet_name = "name"

wb = openpyxl.load_workbook(filename=full_path)
ws = wb[sheet_name]
ic(ws.max_row)
ic(ws.max_column)

column_f = 6
column_h = 8

for row in range(1, ws.max_row):
    # ic(ws.cell(row, column_h).value)
    code = ws.cell(row, column_f).value
    code_bottom = ws.cell(row+1, column_f).value
    value = ws.cell(row, column_h).value
    title = "" if value is None else str(value)
    table_re = re.compile(r"Таблица \d+\.\d+-\d+\.")
    match_result = table_re.match(title)
    if match_result:
        if code_bottom is not None:
            d = ws.cell(row=row, column=column_f, value=code_bottom)
            # ic(ws.cell(row, column_f).value)
        else:
            mess = f"пустой код под таблицей  на строке: {row}"
            ic(mess)
mess = f"пройдено {row} строк. "
ic(mess)


wb.save(output_file)
ic("готово!")

