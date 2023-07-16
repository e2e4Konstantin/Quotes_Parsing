import openpyxl
from openpyxl.styles import (NamedStyle, Font, Border, Side, PatternFill, Alignment)
import os

tab_color = ["0099CC00", "00FFCC00", "00FF9900", "00FF6600", "00666699", "00C0C0C0",] # https://openpyxl.readthedocs.io/en/stable/styles.html

d_color = {
    "table": "0099CC00",
    "quote": "00FFCC00",
    "collection": "00FF9900",
    "attribute": "00FF6600",
    "options": "00666699",
    "header": "00C0C0C0",
}


def not_in_use(filename):
    try:
        os.rename(filename, filename)
        return True
    except IOError as err:
        return False

class ExcelControl:
    def __init__(self, excel_file_name: str = None, sheet_name: str = None):
        self.file_name = excel_file_name
        self.sheet_name = sheet_name
        self.book = None
        self.sheet = None
        self.header_style = None

    def __enter__(self):
        """ Вызывается при старте контекстного менеджера.
        Возвращаемое значение, присваивается переменной x в конце выражения with as x.
        """
        self.open_excel_file()
        return self.sheet

    def __exit__(self, exception_type, exception_value, traceback):
        """ Будет вызван в завершении конструкции with, или в случае возникновения ошибки после нее. """
        self.close_excel_file()

    def __str__(self):
        return f"excel file: {self.file_name}, sheet: {self.sheet_name}, *: {self.sheet}"

    def open_excel_file(self):
        """ Открывает файл, если есть страница self.sheet_nam, удаляет ее и создает новую. """
        try:
            self.book = openpyxl.load_workbook(self.file_name)
            if self.sheet_name in self.book.sheetnames:
                idx = self.book.worksheets.index(self.book[self.sheet_name])
                # print(f"{idx = } {self.book._sheets = }")
                self.book.remove(self.book[self.sheet_name])
            else:
                idx = -1
            self.sheet = self.book.create_sheet(self.sheet_name)
            self.book.active = self.sheet


            if idx >= 0:
                # self.book.move_sheet(self.book[self.sheet_name], 0)
                sheets = self.book._sheets                                  # порядок листов в книге
                sheet = sheets.pop(self.book.worksheets.index(self.sheet))
                sheets.insert(idx, sheet)
            self.sheet.sheet_properties.tabColor = "1072BA"

        except IOError as err:
            print(f"ошибка при открытии файла: {self.file_name}\n{err}")
            self.book = openpyxl.Workbook()
            self.sheet = self.book.active
            self.sheet.title = self.sheet_name
            self.set_header_style()

    def close_excel_file(self):
        if not_in_use(self.file_name):
            if self.book:
                self.book.save(self.file_name)
                self.book.close()
        else:
            print(f"Не могу записать файл {self.file_name},\n он используется в другой программе.")

    def set_header_style(self):
        self.header_style = NamedStyle(name="header_style")
        self.header_style.font = Font(name="Tahoma", bold=False, size=10)
        bd = Side(style='thin', color="000000")
        self.header_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.header_style.fill = PatternFill("solid", fgColor=tab_color[5]) # "DDDDDD"
        self.header_style.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True,
                                                shrink_to_fit=False, indent=0)
        if not ("header_style" in self.book.named_styles):
            self.book.add_named_style(self.header_style)


x = ExcelControl("test.xlsx", "tab1")
with x as mysheet:
    d = mysheet.cell(row=4, column=2, value="Point 10")
    mysheet["A1"] = "Point ****10"

    print('**', d.value)
    d.style = "header_style"

    # cell_range = mysheet['A1':'C2']
    # cell_range.style = "header_style"

    for x in range(len(tab_color)):
        mysheet.cell(row=x+1, column=x+1, value="AAAAAAAAAAA").font = Font(color=tab_color[x])

    for x in range(len(tab_color)):
        mysheet.cell(row=x+10, column=x+1, value="BBBBBBB").fill = PatternFill("solid", fgColor=tab_color[x])

    range = mysheet["B8":"D9"]
    print(range[0])
    print(range[1])
    import random

    for row in mysheet['A1:P1']:
        for cell in row:
            ri = random.randint(0,len(tab_color)-1)
            cell.fill = PatternFill("solid", fgColor=tab_color[ri])

    row = {"A": 22, "B": 33}

    print(row)
    mysheet.append(row)


    mysheet.cell(1, 1, 'World')          # write_string()
    mysheet.cell(2, 1, 2)                # write_number()
    mysheet.cell(3, 1, 3.00001)          # write_number()
    mysheet.cell(4, 1, '=SIN(PI()/4)')   # write_formula()
    mysheet.cell(5, 1, '')               # write_blank()
    mysheet.cell(6, 1, None)             # write_blank()


    for rows in mysheet.iter_rows(min_row=10, max_row=50, min_col=None):
        for cell in rows:
            cell.fill = PatternFill("solid", fgColor=tab_color[random.randint(0,len(tab_color)-1)])

    for row in mysheet['A1':'D4']:
        for cell in row:
            print(cell.coordinate, cell.value)

    from openpyxl.utils import get_column_letter, column_index_from_string

    print(column_index_from_string('F'), get_column_letter(5))