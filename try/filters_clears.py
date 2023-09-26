import openpyxl

wb = openpyxl.load_workbook(filename="data.xlsx")
ws = wb['tab']

# отключает выбор фильтра в заголовке
ws.auto_filter.ref = None


# объекты ws.row_dimensions и ws.column_dimensions ссылки на строку и на столбец
# ws.row_dimensions[2] содержит массив ячеек, которые расположены в строке с номером 2.
# ws.column_dimensions['B'] будет содержать в себе массив ячеек, которые расположены в столбце с буквой 'B'.
#   letter = get_column_letter(i)
#     # получаем ширину столбца
#     col_width = ws.column_dimensions[letter].width
#


# проход по строкам и восстановление скрытых
for i in range(2, ws.max_row + 1):
    ws.row_dimensions[i].hidden = False

wb.save("data.xlsx")


# # set table
# tbl = ws.tables["tablename"]
# # removes all filters
# tbl.autoFilter.filterColumn.clear()

