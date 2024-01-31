
import openpyxl

colors = ["0099CC00", "00FFCC00", "00FF9900", "000066CC", "00666699", "00C0C0C0", "0099CCFF"]
items = ['Quote', 'Attributes', 'Options', 'Collections',  'Tables', 'Header']

tab_color = dict(zip(items, colors))


fp = r"F:\Kazak\Google Диск\1_KK\Job_CNAC\Python_projects\Pars3\template_4_68_output.xlsx"

b = openpyxl.load_workbook(fp)
# wb = openpyxl.Workbook(fp)

for sheet in b.worksheets:
    print(sheet.title)
print(f"{b.sheetnames = }")

for sheet in b.worksheets:
    b.remove(sheet)

for name in tab_color:
    sheet = b.create_sheet(name)
    sheet.sheet_properties.tabColor = tab_color[name]

print(f"{b.sheetnames = }")

if b:
   b.save("1__1.xlsx")
   b.close()
