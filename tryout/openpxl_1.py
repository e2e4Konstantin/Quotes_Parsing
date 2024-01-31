import datetime
import openpyxl
# from openpyxl import Workbook

wb = openpyxl.Workbook()
ws = wb.active
print(ws.title)
ws.title = 'Fruit'
print(ws.title)
print(wb.sheetnames)
sheet1 = wb['Fruit']

ws2 = wb.create_sheet("S2", 0) # insert at first position
print(wb.sheetnames)


ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.datetime.now()

c = ws['A5']
print(c)
ws['A5'] = 4
print(ws['A5'].value)
d = ws.cell(row=4, column=2, value=10)
print(d.value)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

for i, s in enumerate(wb.sheetnames):
    if s == 'charlie':
        break
wb.active = i


wb.save("sample.xlsx")

# idx = len(self.book.worksheets)

# with load_worksheet_with_close('myworkbook.xlsx') as wb:
#     # Do things with workbook
# d = ws.cell(row=4, column=2, value=10)
# cell_range = ws['A1':'C2']
# colC = ws['C']
# col_range = ws['C:D']
# row10 = ws[10]
# row_range = ws[5:10]
# >>> ws = wb.active
# >>> ws['C9'] = 'hello world'
# >>> tuple(ws.rows)
# ((<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>),
# (<Cell Sheet.A2>, <Cell Sheet.B2>, <Cell Sheet.C2>),
# (<Cell Sheet.A3>, <Cell Sheet.B3>, <Cell Sheet.C3>),
# (<Cell Sheet.A4>, <Cell Sheet.B4>, <Cell Sheet.C4>),
# (<Cell Sheet.A5>, <Cell Sheet.B5>, <Cell Sheet.C5>),
# (<Cell Sheet.A6>, <Cell Sheet.B6>, <Cell Sheet.C6>),
# (<Cell Sheet.A7>, <Cell Sheet.B7>, <Cell Sheet.C7>),
# (<Cell Sheet.A8>, <Cell Sheet.B8>, <Cell Sheet.C8>),
# (<Cell Sheet.A9>, <Cell Sheet.B9>, <Cell Sheet.C9>))

for row in mysheet['A1:P1']:
    for cell in row:
        ri = random.randint(0, len(tab_color) - 1)
        cell.fill = PatternFill("solid", fgColor=tab_color[ri])

row = {"A": 22, "B": 33}

print(row)
mysheet.append(row)

mysheet.cell(1, 1, 'World')  # write_string()
mysheet.cell(2, 1, 2)  # write_number()
mysheet.cell(3, 1, 3.00001)  # write_number()
mysheet.cell(4, 1, '=SIN(PI()/4)')  # write_formula()
mysheet.cell(5, 1, '')  # write_blank()
mysheet.cell(6, 1, None)  # write_blank()
