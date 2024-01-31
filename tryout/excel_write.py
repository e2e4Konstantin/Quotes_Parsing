import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from utilites import tables, quotes, collections, Collection, Quote
from dataclasses import fields


def excel_write_console(book_name: str, text_in: str):
    if len(text_in) > 0:
        try:
            book = openpyxl.load_workbook(book_name)
        except IOError as err:
            print(err)
            book = openpyxl.Workbook()
        name = 'Console'
        name_sheets = book.sheetnames
        if name in name_sheets:
            book.remove(book[name])
        book.create_sheet(title=name)
        sheet = book[name]

        sheet['A1'] = text_in
        sheet['A1'].alignment = Alignment(wrapText=True)

        book.save(book_name)
        book.close()


def excel_write_collection(bookname: str):
    if len(collections) > 0 and len(bookname) >= 6:
        try:
            book = openpyxl.load_workbook(bookname)
        except IOError as err:
            print(err)
            book = openpyxl.Workbook()
        name = 'Collections'
        name_sheets = book.sheetnames
        if name in name_sheets:
            book.remove(book[name])
        book.create_sheet(index=1, title=name)
        sheet = book[name]
        head = ['Строка', 'Шифр', 'Наименование сборника', 'Всего расценок', 'Параметризованных', 'Пустых',
                'Коды пустых расценок']
        sheet.append(head)
        for c in range(1, 8):
            sheet.cell(row=1, column=c).fill = PatternFill("solid", fgColor="DDDDDD")
        # line_data
        for collection in collections:
            line_data = [getattr(collections[collection], x.name) for x in fields(Collection)]
            total_quotes = collections[collection].quantity_parameterized_quotes + collections[
                collection].quantity_not_parameterized_quotes
            line_data.insert(3, total_quotes)
            bad_list = line_data.pop()  # удаляем последний элемент
            bad_list_string = ', '.join(x for x in bad_list)
            line_data.append(bad_list_string)
            # print(line_data)
            sheet.append(line_data)
        book.save(bookname)
        book.close()





def excel_write_quotes(book_name: str):
    if len(quotes) > 0 and len(book_name) >= 6:
        try:
            book = openpyxl.load_workbook(book_name)
        except IOError as err:
            print(err)
            book = openpyxl.Workbook()
        name = 'Quotes'
        name_sheets = book.sheetnames
        if name in name_sheets:
            book.remove(book[name])
        book.create_sheet(title=name)
        sheet = book[name]
        head = ["ROW", "GROUP_WORK_PROCESS", "PRESSMARK", "TITLE", "UNIT_OF_MEASURE", "STAT_SUM", "PARAMETERIZED_FLAG",
                "SUPPLEMENTARY_TYPE", "PARENT_PRESSMARK"]
        sheet.append(head)
        for c in range(1, len(head) + 1):
            sheet.cell(row=1, column=c).fill = PatternFill("solid", fgColor="DDDDDD")
        for quote in quotes:
            line_data = [getattr(quote, x.name) for x in fields(Quote)]
            sheet.append(line_data[:-2])

        book.save(book_name)
        book.close()


def excel_write_attributes(book_name: str):
    if len(quotes) > 0 and len(book_name) >= 6:
        try:
            book = openpyxl.load_workbook(book_name)
        except IOError as err:
            print(err)
            book = openpyxl.Workbook()
        name = 'ATTRIBUTES'
        name_sheets = book.sheetnames
        if name in name_sheets:
            book.remove(book[name])
        book.create_sheet(title=name)
        sheet = book[name]
        head = ["ROW", "PRESSMARK", "ATTRIBUTE_TITLE", "VALUE", ]
        sheet.append(head)
        for c in range(1, len(head) + 1):
            sheet.cell(row=1, column=c).fill = PatternFill("solid", fgColor="DDDDDD")
        data_line = list()
        for quote in quotes:
            for attribute in quote.attributes_quote:
                data_line.clear()
                data_line.append(quote.row_quote)
                data_line.append(quote.cod_quote)
                data_line.append(str(attribute.name_attribute).strip().capitalize())
                data_line.append(str(attribute.value_attribute).strip().capitalize())
                sheet.append(data_line)

        book.save(book_name)
        book.close()


def excel_write_options(book_name: str):
    if len(quotes) > 0 and len(book_name) >= 6:
        try:
            book = openpyxl.load_workbook(book_name)
        except IOError as err:
            print(err)
            book = openpyxl.Workbook()
        name = 'PARAMETER'
        name_sheets = book.sheetnames
        if name in name_sheets:
            book.remove(book[name])
        book.create_sheet(title=name)
        sheet = book[name]
        head = ["ROW", "PRESSMARK", "PARAMETER_TITLE", "LEFT_BORDER", "RIGHT_BORDER", "UNIT_OF_MEASURE", "STEP",
                "PARAMETER_TYPE"]

        sheet.append(head)
        for c in range(1, len(head) + 1):
            sheet.cell(row=1, column=c).fill = PatternFill("solid", fgColor="DDDDDD")
        data_line = list()
        for quote in quotes:
            for option in quote.options_quote:
                data_line.clear()
                data_line.append(quote.row_quote)
                data_line.append(quote.cod_quote)
                data_line.append(option.name_option)
                for value in option.value_option:
                    data_line.append(value[1])
                sheet.append(data_line)
        book.save(book_name)
        book.close()


def excel_write_tables(book_name: str):
    if len(tables) > 0 and len(book_name) >= 6:
        try:
            book = openpyxl.load_workbook(book_name)
        except IOError as err:
            print(err)
            book = openpyxl.Workbook()
        name = 'Tables'
        name_sheets = book.sheetnames
        if name in name_sheets:
            book.remove(book[name])
        book.create_sheet(title=name)
        sheet = book[name]
        head = ["row", "номер", "код", "атрибутов", "параметров", "название", "атрибуты", "параметры"]

        sheet.append(head)
        for c in range(1, len(head) + 1):
            sheet.cell(row=1, column=c).fill = PatternFill("solid", fgColor="DDDDDD")
        for table in tables:
            data_line = tables[table].to_list()
            sheet.append(data_line)
        book.save(book_name)
        book.close()

